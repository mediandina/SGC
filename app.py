from flask import Flask, render_template, request, redirect, url_for, jsonify, session, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from datetime import datetime, date
from email.message import EmailMessage

import sqlite3
import os
import re
import smtplib
import io

app = Flask(__name__)
app.secret_key = "clave-secreta-sgf"  # Cambia esto por una clave segura en producción

# -------------------------------------------------------
# CONFIGURACIÓN DE CORREO
# -------------------------------------------------------
EMAIL_REMITENTE    = "multimediandina@gmail.com"
EMAIL_APP_PASSWORD = "tadg fwgb wett wxlj"
# -------------------------------------------------------

DB_FILE = "database.db"

# Credenciales administrador
ADMIN_TEL  = "1234567890"
ADMIN_PASS = "admin123"


# =======================================================
# BASE DE DATOS — INICIALIZACIÓN
# =======================================================
def get_db():
    """Retorna una conexión a SQLite con acceso por nombre de columna."""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Crea las tablas si no existen. Se ejecuta al arrancar."""
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre    TEXT NOT NULL,
                telefono  TEXT NOT NULL UNIQUE,
                proveedor TEXT NOT NULL,
                password  TEXT NOT NULL,
                creado_en TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cupos (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha            TEXT NOT NULL,
                nombre_conductor TEXT NOT NULL,
                tipo_vehiculo    TEXT NOT NULL,
                cupo             INTEGER NOT NULL,
                proveedor        TEXT NOT NULL,
                telefono         TEXT NOT NULL,
                correo           TEXT NOT NULL,
                placa            TEXT NOT NULL,
                kilos            INTEGER NOT NULL,
                pacas            INTEGER NOT NULL,
                registrado_en    TEXT DEFAULT (datetime('now','localtime')),
                UNIQUE(fecha, cupo)
            )
        """)
        conn.commit()


init_db()


# =======================================================
# GENERAR EXCEL EN MEMORIA (para descarga del admin)
# =======================================================
def generar_excel():
    """Consulta la BD y devuelve un BytesIO con el Excel formateado."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT fecha, nombre_conductor, tipo_vehiculo, cupo,
                   proveedor, telefono, correo, placa, kilos, pacas
            FROM cupos
            ORDER BY fecha, cupo
        """).fetchall()

    columnas = [
        "Fecha", "Nombre del conductor", "Tipo de vehículo", "Cupo",
        "Proveedor", "Telefono", "Correo", "Placa",
        "Kilos aproximados", "Pacas"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cupos"

    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    fill_header = PatternFill("solid", fgColor="C7B404")

    for col_idx, nombre_col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=col_idx, value=nombre_col)
        c.font      = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
        c.fill      = fill_header
        ws.column_dimensions[c.column_letter].width = 22

    for row_idx, row in enumerate(rows, 2):
        valores = [
            row["fecha"], row["nombre_conductor"], row["tipo_vehiculo"],
            row["cupo"], row["proveedor"], row["telefono"],
            row["correo"], row["placa"], row["kilos"], row["pacas"]
        ]
        for col_idx, valor in enumerate(valores, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=valor)
            c.border    = borde
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 6:          # Telefono como texto
                c.number_format = "@"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =======================================================
# CORREO DE CONFIRMACIÓN
# =======================================================
def enviar_correo_confirmacion(destinatario, nombre, fecha, cupo):
    try:
        meses = {1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",
                 7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}
        dias  = {0:"lunes",1:"martes",2:"miercoles",3:"jueves",
                 4:"viernes",5:"sabado",6:"domingo"}
        fo = datetime.strptime(fecha, "%Y-%m-%d")
        fecha_legible = f"{dias[fo.weekday()]} {fo.day} de {meses[fo.month]} de {fo.year}"

        asunto = f"Confirmacion cupo {cupo} - {fecha_legible} - Descargue Andina"

        html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f5f5f3;font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f5f5f3;padding:32px 16px;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.08);">
  <tr><td style="background:linear-gradient(135deg,#6bbf2d,#5aa524);padding:32px 40px;text-align:center;">
    <h1 style="margin:0;color:#fff;font-size:26px;font-weight:700;">Descargue Andina</h1>
    <p style="margin:8px 0 0;color:rgba(255,255,255,.88);font-size:14px;">Sistema de Gestion de Cupos</p>
  </td></tr>
  <tr><td style="background:#f0f9e8;padding:20px 40px;text-align:center;border-bottom:3px solid #6bbf2d;">
    <p style="margin:0 0 6px;font-size:36px;">&#10003;</p>
    <h2 style="margin:0 0 4px;color:#1f2937;font-size:20px;">Cupo confirmado exitosamente</h2>
    <p style="margin:0;color:#4b5563;font-size:14px;">Su agendamiento ha sido registrado en el sistema.</p>
  </td></tr>
  <tr><td style="padding:32px 40px;">
    <p style="margin:0 0 20px;color:#1f2937;font-size:15px;">Hola, <strong>{nombre}</strong>. A continuacion encontrara el detalle de su cupo:</p>
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#f9fafb;border-radius:8px;border:1px solid #e5e7eb;">
      <tr><td style="padding:14px 20px;border-bottom:1px solid #e5e7eb;">
        <span style="color:#6b7280;font-size:11px;font-weight:700;text-transform:uppercase;display:block;margin-bottom:4px;">Fecha asignada</span>
        <span style="color:#1f2937;font-size:16px;font-weight:700;text-transform:capitalize;">{fecha_legible}</span>
      </td></tr>
      <tr><td style="padding:14px 20px;border-bottom:1px solid #e5e7eb;">
        <span style="color:#6b7280;font-size:11px;font-weight:700;text-transform:uppercase;display:block;margin-bottom:6px;">Numero de cupo</span>
        <span style="display:inline-block;background:#6bbf2d;color:#fff;font-size:22px;font-weight:800;padding:6px 24px;border-radius:6px;">Cupo {cupo}</span>
      </td></tr>
      <tr><td style="padding:14px 20px;">
        <span style="color:#6b7280;font-size:11px;font-weight:700;text-transform:uppercase;display:block;margin-bottom:4px;">Conductor</span>
        <span style="color:#1f2937;font-size:15px;font-weight:600;">{nombre}</span>
      </td></tr>
    </table>
    <table width="100%" cellpadding="0" cellspacing="0" style="margin-top:20px;border-left:4px solid #c7b404;background:#fffbf0;border-radius:4px;">
      <tr><td style="padding:14px 16px;">
        <p style="margin:0;color:#1f2937;font-size:14px;line-height:1.6;">
          <strong>Recuerde:</strong> el dia del cupo asignado, el personal de Descargue Andina se comunicara <strong>telefonicamente</strong> con usted para indicarle que puede acercarse al area de descarga.
        </p>
      </td></tr>
    </table>
  </td></tr>
  <tr><td style="padding:0 40px 32px;">
    <h3 style="margin:0 0 14px;color:#1f2937;font-size:15px;font-weight:700;border-bottom:2px solid #e5e7eb;padding-bottom:8px;">Reglas importantes</h3>
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr><td style="padding:8px 0;border-bottom:1px solid #f3f4f6;"><span style="color:#6bbf2d;font-weight:700;margin-right:8px;">&#10003;</span><span style="color:#374151;font-size:13.5px;">Todos los datos solicitados deben ser diligenciados correctamente.</span></td></tr>
      <tr><td style="padding:8px 0;border-bottom:1px solid #f3f4f6;"><span style="color:#6bbf2d;font-weight:700;margin-right:8px;">&#10003;</span><span style="color:#374151;font-size:13.5px;">En caso de no poder asistir, comuniquese con el area de despacho con anticipacion.</span></td></tr>
      <tr><td style="padding:8px 0;border-bottom:1px solid #f3f4f6;"><span style="color:#6bbf2d;font-weight:700;margin-right:8px;">&#10003;</span><span style="color:#374151;font-size:13.5px;">Para el ingreso es obligatorio contar con el ticket de bascula autorizada.</span></td></tr>
      <tr><td style="padding:8px 0;"><span style="color:#6bbf2d;font-weight:700;margin-right:8px;">&#10003;</span><span style="color:#374151;font-size:13.5px;">Por favor comprobar que los datos registrados sean iguales a los del formulario.</span></td></tr>
    </table>
  </td></tr>
  <tr><td style="background:#111213;padding:24px 40px;text-align:center;border-radius:0 0 12px 12px;">
    <p style="margin:0 0 6px;color:#fff;font-size:14px;font-weight:600;">Descargue Andina &mdash; Servicio al cliente</p>
    <p style="margin:0;color:#9ca3af;font-size:12px;">+57 313 8893206 &nbsp;|&nbsp;<a href="mailto:mprimas@corrugadosandina.com.co" style="color:#6bbf2d;text-decoration:none;">mprimas@corrugadosandina.com.co</a></p>
    <p style="margin:12px 0 0;color:#6b7280;font-size:11px;">Correo automatico — no responder.</p>
  </td></tr>
</table>
</td></tr>
</table>
</body></html>"""

        msg = EmailMessage()
        msg["Subject"]  = asunto
        msg["From"]     = f"Descargue Andina <{EMAIL_REMITENTE}>"
        msg["To"]       = destinatario
        msg["Reply-To"] = "mprimas@corrugadosandina.com.co"
        msg.set_content(f"Cupo {cupo} confirmado para el {fecha_legible}.\nConductor: {nombre}\nDescargue Andina | +57 313 8893206")
        msg.add_alternative(html, subtype="html")

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.ehlo()
            smtp.login(EMAIL_REMITENTE, EMAIL_APP_PASSWORD)
            smtp.send_message(msg)

        print(f"[EMAIL OK] Confirmacion enviada a {destinatario}")
        return True

    except Exception as e:
        print(f"[EMAIL ERROR] {destinatario}: {e}")
        return False


# =======================================================
# VALIDACIÓN CENTRAL
# =======================================================
def validar_datos(form):
    errores = []

    if not form.get("nombre") or len(form.get("nombre")) > 30:
        errores.append("Nombre invalido")
    if not form.get("proveedor") or len(form.get("proveedor")) > 30:
        errores.append("Proveedor invalido")

    fecha = form.get("fecha", "")
    if not fecha:
        errores.append("Fecha invalida")
    else:
        try:
            datetime.strptime(fecha, "%Y-%m-%d")
        except Exception:
            errores.append("Fecha invalida")

    cupo = form.get("cupo", "")
    if not cupo or not re.fullmatch(r"\d+", str(cupo)):
        errores.append("Cupo invalido")
    else:
        try:
            if not (1 <= int(cupo) <= 100):
                errores.append("Cupo fuera de rango")
        except Exception:
            errores.append("Cupo invalido")

    if not re.fullmatch(r"\d{10}", form.get("telefono", "")):
        errores.append("Telefono invalido")
    if not re.fullmatch(r"\d{10}", session.get("usuario", "")):
        errores.append("Sesion invalida")

    placa = form.get("placa", "").upper().strip()
    if not re.fullmatch(r"[A-Z]{3}[0-9]{3}", placa):
        errores.append("Placa invalida")

    try:
        if not (1 <= int(form.get("kilos")) <= 50000):
            errores.append("Kilos fuera de rango")
    except Exception:
        errores.append("Kilos invalidos")

    try:
        if not (1 <= int(form.get("pacas")) <= 80):
            errores.append("Pacas fuera de rango")
    except Exception:
        errores.append("Pacas invalidas")

    return errores


# =======================================================
# RUTAS
# =======================================================

@app.route("/registro", methods=["GET", "POST"])
def registro():
    if "usuario" in session:
        return redirect(url_for("sistema"))

    if request.method == "POST":
        try:
            telefono = re.sub(r"\D", "", request.form.get("telefono", "").strip())

            with get_db() as conn:
                if conn.execute("SELECT 1 FROM usuarios WHERE telefono=?", (telefono,)).fetchone():
                    return render_template("error.html", mensaje="El telefono ya esta registrado")

                conn.execute(
                    "INSERT INTO usuarios (nombre,telefono,proveedor,password) VALUES (?,?,?,?)",
                    (request.form.get("nombre"), telefono,
                     request.form.get("proveedor"),
                     generate_password_hash(request.form.get("password")))
                )
                conn.commit()

            session["usuario"] = telefono
            return redirect(url_for("sistema"))

        except Exception as e:
            print("ERROR registro:", e)
            return render_template("error.html", mensaje="Ocurrio un error inesperado."), 500

    return render_template("formulario.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if "usuario" in session:
        return redirect(url_for("sistema"))

    if request.method == "POST":
        telefono = re.sub(r"\D", "", request.form.get("telefono", "").strip())
        password = request.form.get("password", "")

        if telefono == ADMIN_TEL and password == ADMIN_PASS:
            session["usuario"]  = telefono
            session["is_admin"] = True
            return redirect(url_for("admin_panel"))

        with get_db() as conn:
            user = conn.execute("SELECT * FROM usuarios WHERE telefono=?", (telefono,)).fetchone()

        if not user:
            return render_template("error.html", mensaje="Usuario no encontrado")
        if not check_password_hash(user["password"], password):
            return render_template("error.html", mensaje="Contrasena incorrecta")

        session["usuario"] = telefono
        session.pop("is_admin", None)
        return redirect(url_for("sistema"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def sistema():
    if "usuario" not in session:
        return redirect(url_for("login"))
    if session.get("is_admin"):
        return redirect(url_for("admin_panel"))
    return render_template("sistema.html")


@app.route("/admin")
def admin_panel():
    if "usuario" not in session or not session.get("is_admin"):
        return redirect(url_for("login"))
    try:
        with get_db() as conn:
            rows = conn.execute("""
                SELECT fecha, nombre_conductor, tipo_vehiculo, cupo,
                       proveedor, telefono, correo, placa, kilos, pacas
                FROM cupos ORDER BY fecha, cupo
            """).fetchall()

        columnas = ["Fecha","Nombre del conductor","Tipo de vehículo","Cupo",
                    "Proveedor","Telefono","Correo","Placa","Kilos aproximados","Pacas"]

        filas = [{
            "Fecha": r["fecha"], "Nombre del conductor": r["nombre_conductor"],
            "Tipo de vehículo": r["tipo_vehiculo"], "Cupo": r["cupo"],
            "Proveedor": r["proveedor"], "Telefono": r["telefono"],
            "Correo": r["correo"], "Placa": r["placa"],
            "Kilos aproximados": r["kilos"], "Pacas": r["pacas"]
        } for r in rows]

        return render_template("admin.html", columnas=columnas, filas=filas)

    except Exception as e:
        print("ERROR admin_panel:", e)
        return render_template("error.html", mensaje="No se pudo cargar los datos"), 500


@app.route("/download_cupos")
def download_cupos():
    if "usuario" not in session or not session.get("is_admin"):
        return redirect(url_for("login"))
    try:
        excel         = generar_excel()
        nombre_archivo = f"cupos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            excel,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print("ERROR download_cupos:", e)
        return render_template("error.html", mensaje="No se pudo generar el Excel"), 500


@app.route("/cupos_ocupados")
def cupos_ocupados():
    if "usuario" not in session:
        return jsonify({"ocupadas": []})
    try:
        fecha    = request.args.get("fecha", "")
        ocupadas = []
        if fecha:
            with get_db() as conn:
                rows     = conn.execute("SELECT cupo FROM cupos WHERE fecha=?", (fecha,)).fetchall()
                ocupadas = [r["cupo"] for r in rows]
        return jsonify({"ocupadas": ocupadas})
    except Exception:
        return jsonify({"ocupadas": []})


@app.route("/guardar", methods=["POST"])
def guardar_cupo():
    if "usuario" not in session:
        return redirect(url_for("login"))
    try:
        errores = validar_datos(request.form)
        if errores:
            return render_template("error.html", mensaje="Datos invalidos enviados al sistema"), 400

        fecha = request.form.get("fecha")
        cupo  = int(request.form.get("cupo"))

        fecha_cupo = datetime.strptime(fecha, "%Y-%m-%d").date()
        hoy        = date.today()

        if fecha_cupo < hoy:
            return render_template("error.html", mensaje="No se pueden registrar cupos para fechas pasadas"), 400
        if fecha_cupo.weekday() == 6:
            return render_template("error.html", mensaje="No se pueden registrar cupos los domingos"), 400

        max_cupos = 12 if fecha_cupo.weekday() <= 4 else 5
        if not (1 <= cupo <= max_cupos):
            return render_template("error.html", mensaje="Numero de cupo no valido para ese dia"), 400

        nombre = request.form.get("nombre")
        correo = request.form.get("email")

        with get_db() as conn:
            if conn.execute("SELECT 1 FROM cupos WHERE fecha=? AND cupo=?", (fecha, cupo)).fetchone():
                return render_template("error.html", mensaje="El cupo seleccionado ya esta ocupado"), 409

            conn.execute("""
                INSERT INTO cupos
                    (fecha,nombre_conductor,tipo_vehiculo,cupo,proveedor,
                     telefono,correo,placa,kilos,pacas)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                fecha, nombre, request.form.get("tipo"), cupo,
                request.form.get("proveedor"), str(request.form.get("telefono")),
                correo, request.form.get("placa","").upper().strip(),
                int(request.form.get("kilos")), int(request.form.get("pacas"))
            ))
            conn.commit()

        enviar_correo_confirmacion(destinatario=correo, nombre=nombre, fecha=fecha, cupo=cupo)

        return redirect(url_for("sistema", success=1, cupo=cupo, fecha=fecha))

    except Exception as e:
        print("ERROR guardar_cupo:", e)
        return render_template("error.html", mensaje="Ocurrio un error inesperado."), 500


@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private"
    response.headers["Pragma"]        = "no-cache"
    response.headers["Expires"]       = "0"
    return response


@app.errorhandler(400)
def error_400(e): return render_template("error.html", mensaje="Solicitud invalida"), 400
@app.errorhandler(403)
def error_403(e): return render_template("error.html", mensaje="Acceso prohibido"), 403
@app.errorhandler(404)
def error_404(e): return render_template("error.html", mensaje="Pagina no encontrada"), 404
@app.errorhandler(409)
def error_409(e): return render_template("error.html", mensaje="Conflicto en la solicitud"), 409
@app.errorhandler(500)
def error_500(e): return render_template("error.html", mensaje="Error interno del servidor"), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)